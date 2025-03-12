import openpyxl

def getRowCount(filePath):

   workbook= openpyxl.load_workbook(filePath)
   sheet= workbook.active
   return sheet.max_row

def getColumnCount(filePath):

   workbook= openpyxl.load_workbook(filePath)
   sheet= workbook.active
   return sheet.max_column

def readData(filePath, row, col):
    workbook = openpyxl.load_workbook(filePath)
    sheet = workbook.active
    return sheet.cell(row, col).value

def writeData(filePath, row, col,data):

    workbook = openpyxl.load_workbook(filePath)
    sheet = workbook.active
    sheet.cell(row, col).value=data
    workbook.save(filePath)