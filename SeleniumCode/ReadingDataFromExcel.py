import openpyxl
# reading the excel sheet data

filepath="C:\\Users\\JAYA VYAWHARE\\Documents\\BookExcel.xlsx"

workBook=openpyxl.load_workbook(filepath)

sheet=workBook.active

#workBook.get_index(0)
#workBook.get_sheet_by_name("Sheet3")

rowCount=sheet.max_row
print(rowCount)

colCount=sheet.max_column
print(colCount)

print(sheet.cell(8,1).value)

"""
for r in range(1,rowCount+1):
    for c in range(1,colCount+1):

        print(sheet.cell(r,c).value, end="                      ")

    print()


# writing the data into excel sheet

filepath= "C:\\Users\\JAYA VYAWHARE\\Documents\\Book1.xlsx"

workbook=openpyxl.load_workbook(filepath)
sheet=workbook.active

for r in range(1,3):
    for c in range(1, 3):
       sheet.cell(r,c).value="rahul"

workbook.save(filepath)


sheet.cell(5,5).value="Nilesh"

"""